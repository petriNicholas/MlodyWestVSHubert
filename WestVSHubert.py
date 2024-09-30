import requests
import time
import openpyxl

from openpyxl.chart import Reference
from openpyxl.chart import BarChart
from bs4 import BeautifulSoup

hubert = requests.get("https://open.spotify.com/artist/6Vhpr2gUnE0VcGjzqDmM62")
west = requests.get("https://open.spotify.com/artist/1bOTP9P3CS97UwhBm2WekK")

soupH = BeautifulSoup(hubert.content, 'html.parser')
soupW = BeautifulSoup(west.content, 'html.parser')

hubertListeners = int(soupH.get_text().split("Hubert.")[2].split(" monthly listeners")[0].replace(',',''))
westListeners = int(soupW.get_text().split("West")[2].split(" monthly listeners")[0].replace(',',''))

print("Hubert. : ", hubertListeners)
print("West: ", westListeners)

diffListeners = westListeners - hubertListeners

wb = openpyxl.load_workbook('WestVSHubertBAZA.xlsx')
ws = wb['Liczby']

if (ws.cell(ws.max_row, 2).value != time.gmtime().tm_year &
	ws.cell(ws.max_row, 3).value != time.gmtime().tm_mon  &
	ws.cell(ws.max_row, 4).value != time.gmtime().tm_mday):

	new_row = (ws.max_row+2137, time.gmtime().tm_year, time.gmtime().tm_mon, time.gmtime().tm_mday, westListeners, hubertListeners, diffListeners)

	ws.append(new_row)
else:
	print("no need to update the database")

values = Reference(ws,
                   min_col=5,
                   max_col=6,
                   min_row=1,
                   max_row=ws.max_row)

cats = Reference(ws,
                   min_col=3,
                   max_col=3,
                   min_row=2,
                   max_row=ws.max_row)

chart = BarChart()
chart.add_data(values, titles_from_data=True)
chart.set_categories(cats)

chart.title = "kupa"
chart.x_axis.title = "Miesiąc"
chart.y_axis.title = "Monthly listeners"

ws = wb['Chart']
ws.add_chart(chart,"B2")

wb.save('WestVSHubertBAZA.xlsx')